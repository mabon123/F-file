import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import warnings
from ttkthemes import ThemedTk
import os

# Suppress warnings for unsupported extensions
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class ExcelEditor:

    def __init__(self, file_path):
        """Initialize the editor with the Excel file path."""
        self.file_path = file_path
        self.workbook = None

    def load_workbook(self):
        """Load the workbook from the specified file path."""
        try:
            self.workbook = load_workbook(self.file_path)
            return "ឯកសារបានបញ្ចូលរួចរាល់"
        except Exception as e:
            return f"Error loading workbook: {e}"

    def count_rows_until_condition(self, sheet_name, start_row, column, condition):
        """
        Count rows from a starting row until a condition is met in the specified column.
        Returns the count and whether the condition was met.
        """
        if sheet_name not in self.workbook.sheetnames:
            return f"**Sheet '{sheet_name}' រកមិនឃើញ.", 0, False

        ws = self.workbook[sheet_name]
        count = 0

        for row_num in range(start_row, ws.max_row + 1):
            cell_value = ws[f"{column}{row_num}"].value
            count += 1
            if isinstance(cell_value, str) and cell_value.startswith(condition):
                return "", count, True

        return f"មិនបានជួបលក់ខណ្ឌចំណុច {condition} '{sheet_name}'.", count, False

    def check_row(self, sheet, start_row, column, condition, num_row, name_blog_row):
        count_message, count, condition_met = self.count_rows_until_condition(sheet, start_row, column, condition)
        if count_message:
            return count_message,0 # Return the error message if there is one

        if condition_met and count != num_row:
            if count < num_row:
                result = num_row-count
                return f"**Sheet '{sheet}' បានលុប Row ចំនួន {result} នៅត្រងចំណុច '{name_blog_row}'.",result
            elif count > num_row:
                result = num_row-count
                return f"**Sheet '{sheet}' បានបន្ថែម Row ចំនួន {abs(result)} នៅត្រងចំណុច '{name_blog_row}'.",result

        # Return None if the row count is correct and there's no action needed
        return None,0
    
    # Check Validated Grade
    def validate_grade(self,grade_value, student_value,note):
        grade = {
            "ទាប",
            "មធ្យម",
            "ខ្ពស់",
            "ចម្រុះ",
            "ទី១",
            "ទី២",
            "ទី៣",
            "ទី៤",
            "ទី៥",
            "ទី៦"
        }
        if grade_value in grade:
            if student_value is None or student_value <=0 or not isinstance(student_value, (int, float)):
                return f"**សូមបញ្ចូលចំនួនសិស្សឱ្យបានត្រឹមត្រូវ: ទិន្នន័យ '{grade_value}' មិនត្រឹមត្រូវ ក្នុង {note}", 0,student_value
        elif grade_value is not None:
            if student_value is None or student_value <=0 or not isinstance(student_value, (int, float)):
                return f"**សូមបញ្ចូលថ្នាក់រៀន និងចំនួនសិស្សឱ្យបានត្រឹមត្រូវ: ទិន្នន័យ '{grade_value} | {student_value}' មិនត្រឹមត្រូវ ក្នុង {note}", grade_value,student_value
            return f"**សូមបញ្ចូលថ្នាក់រៀនឱ្យបានត្រឹមត្រូវ: ទិន្នន័យ '{grade_value}' គ្មានចំនួនសិស្ស", grade_value,0
        elif grade_value is None:
            if student_value is not None:
                return f"**សូមបញ្ចូលថ្នាក់រៀនអោយបានត្រឹមត្រូវ: គ្មានទិន្នន័យថ្នាក់រៀន ក្នុង {note}", grade_value,student_value
        return None, 0, 0
            
    def validate_levels(self, sheet_name, start_row, end_row, level_col, salary_col, certificate_col, day_col, month_col, year_col, gender_col, position_col, grade_col_1, student_col_1, grade_col_2, student_col_2,grade_col_3, student_col_3,grade_col_4, student_col_4, subject_1_col,add_teach_col, name_blog_row):
        """
        Validate that values in level_col (e.g., 'M') belong to level_salary_a
        and their corresponding values in salary_col (e.g., 'P') exist in the level's set.
        Highlight invalid cells in red.
        """
        # Define a red fill pattern
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        # Define the valid values for each column
        level_salary_a = {
            "ឧត្តម": {
                "ក.3.1", "ក.3.2", "ក.3.3", "ក.3.4",
                "ក.2.1", "ក.2.2", "ក.2.3", "ក.2.4",
                "ក.1.1", "ក.1.2", "ក.1.3", "ក.1.4",
                "ក.1.5", "ក.1.6",
            },
            "មូលដ្ឋាន": {
                "ខ.3.1", "ខ.3.2", "ខ.3.3", "ខ.3.4",
                "ខ.2.1", "ខ.2.2", "ខ.2.3", "ខ.2.4",
                "ខ.1.1", "ខ.1.2", "ខ.1.3", "ខ.1.4",
                "ខ.1.5", "ខ.1.6",
            },
            "បឋម": {
                "គ.1", "គ.2", "គ.3", "គ.4",
                "គ.5", "គ.6", "គ.7", "គ.8",
                "គ.9", "គ.10",
            },
            "មត្តេយ្យ": {
                "គ.1", "គ.2", "គ.3", "គ.4",
                "គ.5", "គ.6", "គ.7", "គ.8",
                "គ.9", "គ.10",
            },
            

        }
        # Add Teacher
        add_teacher = {
            "កិច្ចសន្យា",
            "គ្រូខ្ចី",
            "មន្ត្រីកិច្ចសន្យា",
        }
        # Add Certificate
        certificate = {
            "បណ្ឌិត",
            "អនុបណ្ឌិត",
            "បរិញ្ញាបត្រ",
            "មទភ",
            "មបភ",
            "ក្រោម មបភ"
        }
        # Add Gander
        gander = {
            "ប្រុស",
            "ស្រី"
        }
        # Add Position
        position = {
            "នាយក",
            "នាយករង",
            "លេខាធិការ",
            "បណ្ណារក្ស",
            "បេឡា",
            "គណនេយ្យ",
            "រោងជាង",
            "ទទួលបន្ទុកយុវជន",
            "បរិវច្ឆការី ",
            "ឆ្មាំ",
            "បម្រើការនៅអង្កការ",
            "បន្តការសិក្សា",
            "កំពុងស្នើលុបឈ្មោះ",
            "សុំចូលនិវត្តមុនអាយុ",
            "ទំនេរគ្មានបៀវត្ស",
            "ក្រៅក្របខណ្ឌដើម",
            "បាត់បង់សម្បទាវិជ្ជាជីវៈ",
            "មានជំងឺរ៉ាំរ៉ៃ",
            "លំហែមាតុភាព",
            "បង្រៀន",
        }
        # Add Subject
        subject_1 = {
            "ភាសាខ្មែរ",
            "គណិតវិទ្យា",
            "ភាសាអង់គ្លេស",
            "ភាសាបារាំង",
            "កីឡា",
            "រូបវិទ្យា",
            "គីមីវិទ្យា",
            "ជីវវិទ្យា",
            "ផែនដីវិទ្យា",
            "ប្រវត្តិវិទ្យា",
            "ភូមិវិទ្យា",
            "សីលធម៌-ពលរដ្ឋ",
            "គេហវិទ្យា",
            "សេដ្ឋកិច្ច",
            "ព័ត៌មានវិទ្យា",
            "កសិកម្ម",
            "សិល្បៈ",
            "ដូរ្យតន្រ្តី",
            "នាដសាស្រ្ត",
            "រោងជាង",
            "គ្រប់គ្រងទូទៅ",
            "គ្រប់គ្រងអប់រំ",
            "អេឡិចត្រនិច",
            "អគ្គិសនី",
            "មេកានិច",
            "ភាសារុស្សី",
        }
        # Add Teacher
        add_teach = {
            "២វេន",
            "គួប២កម្រិត",
            "គួប៣កម្រិត",
            "២វេនគួប២កម្រិត",
            "២វេនគួប៣កម្រិត",
        }    

        if sheet_name not in self.workbook.sheetnames:
            return f"Sheet '{sheet_name}' not found."

        ws = self.workbook[sheet_name]
        invalid_entries = []

        for row_num in range(start_row, end_row + 1):
            #level value
            level_value = ws[f"{level_col}{row_num}"].value
            #salary value
            salary_value = ws[f"{salary_col}{row_num}"].value
            #certificate value
            certificate_value = ws[f"{certificate_col}{row_num}"].value
            #day value
            day_value = ws[f"{day_col}{row_num}"].value
            #month value
            month_value = ws[f"{month_col}{row_num}"].value
            #year value
            year_value = ws[f"{year_col}{row_num}"].value
            #gander value
            gander_value = ws[f"{gender_col}{row_num}"].value
            #prosition value
            position_value = ws[f"{position_col}{row_num}"].value
            ## Grade 1
            grade_value_1 = ws[f"{grade_col_1}{row_num}"].value
            student_value_1 = ws[f"{student_col_1}{row_num}"].value
            ## Grade 2
            grade_value_2 = ws[f"{grade_col_2}{row_num}"].value
            student_value_2 = ws[f"{student_col_2}{row_num}"].value
            ## Grade 3
            grade_value_3 = ws[f"{grade_col_3}{row_num}"].value
            student_value_3 = ws[f"{student_col_3}{row_num}"].value
            ## Grade 4
            grade_value_4 = ws[f"{grade_col_4}{row_num}"].value
            student_value_4 = ws[f"{student_col_4}{row_num}"].value
            #subject value
            subject_1_value = ws[f"{subject_1_col}{row_num}"].value
            #add teacher value
            add_teach_value = ws[f"{add_teach_col}{row_num}"].value
            #comment message
            results = []
            # Check Level of Teacher
            if level_value in level_salary_a or gander_value in gander or position_value in position:
                # Check Salary Level if គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...
                if name_blog_row == "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...":
                    if level_value not in add_teacher:
                        results.append(f"**បញ្ចូលខុសក្របខ័ណ្ឌអោយបានត្រឹមត្រូវ '{level_value}' ក្នុង {name_blog_row}")
                        ws[f"{level_col}{row_num}"].fill = red_fill  # Highlight level cell in red
                        if salary_value is not None:
                            results.append(f"**កាំប្រាក់ពុំត្រឹមត្រូវ '{salary_value}' សម្រាប់ក្របខ័ណ្ឌ '{level_value}' ក្នុង {name_blog_row}")
                            ws[f"{salary_col}{row_num}"].fill = red_fill  # Highlight salary cell in red
                            
                # Check Level of Teacher
                elif level_value not in level_salary_a:
                    results.append(f"**បញ្ចូលខុសក្របខ័ណ្ឌអោយបានត្រឹមត្រូវ '{level_value}' ក្នុង {name_blog_row}")
                    ws[f"{level_col}{row_num}"].fill = red_fill  # Highlight level cell in red
                    ws[f"{salary_col}{row_num}"].fill = red_fill
                    
                # Check Salary Level
                elif level_value in level_salary_a:
                    if salary_value not in level_salary_a[level_value]:
                        results.append(f"**កាំប្រាក់ពុំត្រឹមត្រូវ '{salary_value}' សម្រាប់ក្របខ័ណ្ឌ '{level_value}' ក្នុង {name_blog_row}")
                        ws[f"{salary_col}{row_num}"].fill = red_fill  # Highlight salary cell in red
                    elif salary_value is None:
                        results.append(f"**សូមបញ្ចូលកាំប្រាក់អោយបានត្រឹមត្រូវ {level_value} ក្នុង {name_blog_row}")
                        ws[f"{salary_col}{row_num}"].fill = red_fill  # Highlight salary cell in red

                # Check Certificate
                if certificate_value not in certificate:
                    results.append(f"**បញ្ចូលសញ្ញាបត្រពុំត្រឹមត្រូវ {certificate_value} ក្នុង {name_blog_row}")
                    ws[f"{certificate_col}{row_num}"].fill = red_fill  # Highlight certificate cell in red

                # Try convert and check date of birth
                try:
                    day_value = int(day_value)
                    month_value = int(month_value)
                    year_value = int(year_value)
                    # Check Day of month
                    if day_value is None or not isinstance(day_value, int) or day_value < 1 or day_value > 31:
                        results.append(f"**បញ្ចូលថ្ងៃទីពុំត្រឹមត្រូវ ឬពុំបានបញ្ចូលក្នុង {name_blog_row}")
                        ws[f"{day_col}{row_num}"].fill = red_fill  # Highlight day cell in red
                    # Check Month of Year
                    if month_value is None or not isinstance(month_value, int) or month_value < 1 or month_value > 12:
                        results.append(f"**បញ្ចូលខែពុំត្រឹមត្រូវ ឬពុំបានបញ្ចូលក្នុង {name_blog_row}")
                        ws[f"{month_col}{row_num}"].fill = red_fill  # Highlight month cell in red
                    # Check Year
                    if year_value is None or not isinstance(year_value, int) or year_value < 1964 or year_value > 2006:
                        results.append(f"**បញ្ចូលឆ្នាំពុំត្រឹមត្រូវ ឬពុំបានបញ្ចូលក្នុង {year_value} ក្នុង {name_blog_row}")
                        ws[f"{year_col}{row_num}"].fill = red_fill  # Highlight year cell in red

                except (ValueError, TypeError):
                    results.append(f"**ថ្ងៃខែឆ្នាំកំណើត '{day_value}'| {month_value} | {year_value} មិនមែនជាលេខសូមបញ្ចូលអោយបានត្រឹមត្រូវ.")
                    ws[f"{day_col}{row_num}"].fill = red_fill  # Highlight day cell in red
                    ws[f"{month_col}{row_num}"].fill = red_fill  # Highlight month cell in red
                    ws[f"{year_col}{row_num}"].fill = red_fill  # Highlight year cell in red

                # Check Gender
                if gander_value not in gander:
                    results.append(f"**ភេទមិនត្រឹមត្រូវ ឬពុំបានបញ្ចូល {gander_value} ក្នុង {name_blog_row}")
                    ws[f"{gender_col}{row_num}"].fill = red_fill  # Highlight gender cell in red

                # Check Position
                if position_value not in position:
                    results.append(f"**មុខតំណែងមិនត្រឹមត្រូវ ឬពុំបានបញ្ចូល: {position_value} ក្នុង {name_blog_row}")
                    ws[f"{position_col}{row_num}"].fill = red_fill  # Highlight position cell in red

                # Check Grade column 1
                grade_message, grade_value, student_value = self.validate_grade(grade_value_1, student_value_1, "ចំណុចបង្រៀន")
                if grade_message is not None:
                    if grade_value != 0 and student_value != 0:
                        results.append(f"{grade_message}")
                        ws[f"{grade_col_1}{row_num}"].fill = red_fill
                        ws[f"{student_col_1}{row_num}"].fill = red_fill
                    elif grade_value == 0 and student_value != 0:
                        results.append(f"{grade_message}")
                        ws[f"{student_col_1}{row_num}"].fill = red_fill
                    elif grade_value != 0 and student_value == 0:
                        results.append(f"{grade_message}")
                        ws[f"{grade_col_1}{row_num}"].fill = red_fill
                
                # Check Grade Column 2
                grade_message, grade_value, student_value = self.validate_grade(grade_value_2, student_value_2, "ថ្នាក់បង្រៀនបន្ថែម")
                if grade_message is not None:
                    if grade_value != 0 and student_value != 0:
                        results.append(f"{grade_message}")
                        ws[f"{grade_col_2}{row_num}"].fill = red_fill
                        ws[f"{student_col_2}{row_num}"].fill = red_fill
                    elif grade_value == 0 and student_value != 0:
                        results.append(f"{grade_message}")
                        ws[f"{student_col_2}{row_num}"].fill = red_fill
                    elif grade_value != 0 and student_value == 0:
                        results.append(f"{grade_message}")
                        ws[f"{grade_col_2}{row_num}"].fill = red_fill    
                        
                # Check Grade Column 3
                grade_message, grade_value, student_value = self.validate_grade(grade_value_3, student_value_3,"ថ្នាក់បង្រៀនបន្ថែម")
                if grade_message is not None:
                    if grade_value != 0 and student_value != 0:
                        results.append(f"{grade_message}")
                        ws[f"{grade_col_3}{row_num}"].fill = red_fill
                        ws[f"{student_col_3}{row_num}"].fill = red_fill
                    elif grade_value == 0 and student_value != 0:
                        results.append(f"{grade_message}")
                        ws[f"{student_col_3}{row_num}"].fill = red_fill
                    elif grade_value != 0 and student_value == 0:
                        results.append(f"{grade_message}")
                        ws[f"{grade_col_3}{row_num}"].fill = red_fill  
                        
                # Check Grade Column 4
                grade_message, grade_value, student_value = self.validate_grade(grade_value_4, student_value_4,"ថ្នាក់បង្រៀនបន្ថែម")
                if grade_message is not None:
                    if grade_value != 0 and student_value != 0:
                        results.append(f"{grade_message}")
                        ws[f"{grade_col_4}{row_num}"].fill = red_fill
                        ws[f"{student_col_4}{row_num}"].fill = red_fill
                    elif grade_value == 0 and student_value != 0:
                        results.append(f"{grade_message}")
                        ws[f"{student_col_4}{row_num}"].fill = red_fill
                    elif grade_value != 0 and student_value == 0:
                        results.append(f"{grade_message}")
                        ws[f"{grade_col_4}{row_num}"].fill = red_fill  
                    
                # Check Subject if not បឋម and មត្តេយ្យ
                if name_blog_row != "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...":
                    if level_value != "បឋម" and level_value != "មត្តេយ្យ":
                        if subject_1_value not in subject_1:
                                results.append(f"** សូមបញ្ចូលមុខវិជ្ជាឱ្យបានត្រឹមត្រូវ {subject_1_value} ក្នុងក្របខ័ណ្ឌ {level_value} នៃ {name_blog_row}")
                                ws[f"{subject_1_col}{row_num}"].fill = red_fill  # Highlight subject cell in red
                
                # Check Add Type Teacher
                if add_teach_value not in add_teach and add_teach_value is not None:
                    results.append(f"** សូមបញ្ចូលប្រភេទនៃការបង្រៀនបន្ថែមរបស់គ្រូបង្រៀនឱ្យបានត្រឹមត្រូវ ក្នុង {name_blog_row}")
                
            if results:
                results.insert(0, f"##### សូមពិនិត្យ Row {row_num} : #####")
                invalid_entries.append(results)

        if invalid_entries:
            # Flatten the list and ensure all entries are strings
            flattened_entries = [item for sublist in invalid_entries for item in sublist]
            return f"********ការផ្ទៀតផ្ទាត់នៅក្នុងបំណែងចែកភារកិច្ចមានបញ្ហា********** '{sheet_name}':\n" + "\n".join(map(str, flattened_entries))
    
    def update_cell(self, sheet_name, cell, value):
        """Update a specific cell with a new value."""
        update_check = []
        if sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            ws[cell] = f"{value}"
        else:
            update_check.append(f"**Sheet '{sheet_name}' រកមិនឃើញ.")

        if update_check:
            return f"**Sheet រកពុំឃើញ '{sheet_name}':\n" + "\n".join(update_check)

    def process_sheets(self):
        """Process all sheets starting with 'S' 
        and within the numeric range."""
        
        results = []
        for sheet in self.workbook.sheetnames:
            if sheet.startswith('S') and sheet[1:].isdigit():
                results.append(f"ដំណើរការផ្ទៀងផ្ទាត់នៅក្នុង: {sheet}")

                # Check rows for different categories
                count = 50
                admin_message = None
                admin_message,result_admin_row = self.check_row(sheet, count, "A", "ខ", 101, "ក.បុគ្គលិកទីចាត់ការ")
                ## Check Admin Row
                if result_admin_row > 0:
                    results.append(admin_message)
                    results.append(self.validate_levels(sheet, count, (149-result_admin_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "ក.បុគ្គលិកទីចាត់ការ"))
                    count =151-result_admin_row
                    teacher_message,result_teacher_row = self.check_row(sheet, count, "A", "គ", 151, "ខ. គ្រូបង្រៀនបឋមសិក្សា")
                    
                    ## Check Teacher Row
                    if result_teacher_row > 0:
                        results.append(teacher_message)
                        results.append(self.validate_levels(sheet, count, (300-result_admin_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "ខ.បំណែងចែកភារកិច្ចគ្រូបង្រៀន"))
                        count = 302 - result_teacher_row - result_admin_row
                        contract_message, result_contract_row = self.check_row(sheet, count, "A", "សរុប", 51, "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...")
                        ## Check Contract Row
                        if contract_message:
                            results.append(contract_message)
                        results.append(self.validate_levels(sheet, count, (351-result_admin_row-result_teacher_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា..."))
                            
                    elif result_teacher_row < 0:
                        results.append(teacher_message)
                        results.append(self.validate_levels(sheet, count, (300-result_admin_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "ខ.បំណែងចែកភារកិច្ចគ្រូបង្រៀន"))
                        count = 302 + abs(result_teacher_row) - result_admin_row
                        contract_message, result_contract_row = self.check_row(sheet, count, "A", "សរុប", 51, "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...")
                        ## Check Contract Row
                        if contract_message:
                            results.append(contract_message)
                        results.append(self.validate_levels(sheet, count, (351-result_admin_row-result_teacher_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា..."))
                            
                    else:
                        results.append(self.validate_levels(sheet, count, (300-result_admin_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "ខ.បំណែងចែកភារកិច្ចគ្រូបង្រៀន"))
                        count = 302-result_admin_row
                        contract_message, result_contract_row = self.check_row(sheet, count, "A", "សរុប", 51, "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...")
                        ## Check Contract Row
                        if contract_message:
                            results.append(contract_message)
                        results.append(self.validate_levels(sheet, count, (351-result_admin_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា..."))
                            
                elif result_admin_row < 0:
                    results.append(admin_message)
                    results.append(self.validate_levels(sheet, count, (149-result_admin_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "ក.បុគ្គលិកទីចាត់ការ"))
                    count = 151 + abs(result_admin_row)
                    teacher_message,result_teacher_row = self.check_row(sheet, count, "A", "គ", 151, "ខ. គ្រូបង្រៀនបឋមសិក្សា")
                    ## Check Teacher Row
                    if result_teacher_row > 0:
                        results.append(teacher_message)
                        results.append(self.validate_levels(sheet, count, (300-result_admin_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "ខ.បំណែងចែកភារកិច្ចគ្រូបង្រៀន"))
                        count = 302 - result_teacher_row +abs(result_admin_row)
                        contract_message, result_contract_row = self.check_row(sheet, count, "A", "សរុប", 51, "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...")
                        ## Check Contract Row
                        if contract_message:
                            results.append(contract_message)
                        results.append(self.validate_levels(sheet, count, (351-result_admin_row-result_teacher_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា..."))
                            
                    elif result_teacher_row < 0:
                        results.append(teacher_message)
                        results.append(self.validate_levels(sheet, count, (300-result_admin_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "ខ.បំណែងចែកភារកិច្ចគ្រូបង្រៀន"))
                        count = 302 + abs(result_teacher_row) + abs(result_admin_row)
                        contract_message, result_contract_row = self.check_row(sheet, count, "A", "សរុប", 51, "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...")
                        ## Check Contract Row
                        if contract_message:
                            results.append(contract_message)
                        results.append(self.validate_levels(sheet, count, (351-result_admin_row-result_teacher_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា..."))
                            
                    else:
                        
                        results.append(self.validate_levels(sheet, count, (300-result_admin_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC","AD", "AF","AG","O","AL", "ខ.បំណែងចែកភារកិច្ចគ្រូបង្រៀន"))
                        count = 302+result_admin_row
                        contract_message, result_contract_row = self.check_row(sheet, count, "A", "សរុប", 51, "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...")
                        ## Check Contract Row
                        if contract_message:
                            results.append(contract_message)
                        results.append(self.validate_levels(sheet, count, (351-result_admin_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា..."))
                            
                else:
                    results.append(self.validate_levels(sheet, count, 149, "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "ក.បុគ្គលិកទីចាត់ការ"))
                    count = 151
                    teacher_message,result_teacher_row = self.check_row(sheet, count, "A", "គ", 151, "ខ. គ្រូបង្រៀនបឋមសិក្សា")
                    ## Check Teacher Row
                    if result_teacher_row > 0:
                        results.append(teacher_message)
                        results.append(self.validate_levels(sheet, count, 300, "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "ខ.បំណែងចែកភារកិច្ចគ្រូបង្រៀន"))
                        count = 302 - result_teacher_row
                        contract_message, result_contract_row = self.check_row(sheet, count, "A", "សរុប", 51, "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...")
                        ## Check Contract Row
                        if contract_message:
                            results.append(contract_message)
                        results.append(self.validate_levels(sheet, count, (351-result_teacher_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា..."))
                            
                    elif result_teacher_row < 0:
                        count = 302 + abs(result_teacher_row)
                        contract_message, result_contract_row = self.check_row(sheet, count, "A", "សរុប", 51, "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...")
                        ## Check Contract Row
                        if contract_message:
                            results.append(contract_message)
                        results.append(self.validate_levels(sheet, count, (351-result_teacher_row), "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា..."))
                    else:
                        results.append(self.validate_levels(sheet, count, 300, "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "ខ.បំណែងចែកភារកិច្ចគ្រូបង្រៀន"))
                        count = 302
                        contract_message, result_contract_row = self.check_row(sheet, count, "A", "សរុប", 51, "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា...")
                        ## Check Contract Row
                        if contract_message:
                            results.append(contract_message)
                        results.append(self.validate_levels(sheet, count, 351, "K", "L", "S", "G", "H", "I", "J", "M","W","X", "Z", "AA", "AC", "AD","AF","AG","O","AL", "គ.គ្រូបង្រៀនជាប់កិច្ចសន្យា..."))
            else:
                results.append(f"មិនមែនជា Sheet សាលារំលង: {sheet}")

        # Filter out None values before joining
        return "\n".join(filter(None, results))

    def save_workbook(self):
        """Save the workbook back to the file."""
        try:
            self.workbook.save(self.file_path)
            return "កិច្ចការបានរក្សាទុករួចរាល់។"
        except Exception as e:
            return f"មានបញ្ហាក្នុងការរក្សារទុក ឬ Excel កំពុងបើក: {e}"



class ModernExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Processor")
        self.root.geometry("800x700")
        self.editor = None
        
        # Set the Khmer font and styles
        self.khmer_font = ("Khmer OS Siemreap", 11)
        self.khmer_font_bold = ("Khmer OS Siemreap", 11, "bold")
        
        # Configure the theme
        self.configure_styles()
        
        # Create UI elements
        self.create_widgets()
        
    def configure_styles(self):
        # Configure custom styles
        style = ttk.Style()
        
        # Configure frame styles
        style.configure("Custom.TFrame", background="#f0f0f0")
        
        # Configure button styles
        style.configure("Primary.TButton",
                       font=self.khmer_font,
                       padding=10)
        
        # Configure label styles
        style.configure("Header.TLabel",
                       font=self.khmer_font_bold,
                       padding=5)
        
        style.configure("Info.TLabel",
                       font=self.khmer_font,
                       padding=5)
        
        # Configure labelframe styles
        style.configure("Custom.TLabelframe",
                       font=self.khmer_font_bold,
                       padding=10)
        style.configure("Custom.TLabelframe.Label",
                       font=self.khmer_font_bold)

    def create_widgets(self):
        # Main container with padding
        main_container = ttk.Frame(self.root, style="Custom.TFrame", padding="20")
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header_frame = ttk.Frame(main_container)
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        header_label = ttk.Label(header_frame,
                                text="កម្មវិធីត្រួតពិនិត្យទិន្នន័យគ្រូបង្រៀន",
                                style="Header.TLabel",
                                font=("Khmer OS Siemreap", 16, "bold"))
        header_label.pack()
        
        # File Selection Frame
        file_frame = ttk.LabelFrame(main_container,
                                  text="ជ្រើសរើសឯកសារ",
                                  style="Custom.TLabelframe",
                                  padding="15")
        file_frame.pack(fill=tk.X, pady=(0, 20))
        
        # File info container
        file_info_frame = ttk.Frame(file_frame)
        file_info_frame.pack(fill=tk.X, pady=(0, 10))
        
        # File icon label (you can replace with actual icon)
        file_icon_label = ttk.Label(file_info_frame, text="📄", font=("", 24))
        file_icon_label.pack(side=tk.LEFT, padx=(0, 10))
        
        self.file_label = ttk.Label(file_info_frame,
                                  text="ឯកសារមិនទាន់បានជ្រើសរើស",
                                  style="Info.TLabel",
                                  wraplength=600)
        self.file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Buttons container
        buttons_frame = ttk.Frame(file_frame)
        buttons_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.select_button = ttk.Button(buttons_frame,
                                      text="ជ្រើសរើសឯកសារ",
                                      style="Primary.TButton",
                                      command=self.select_file)
        self.select_button.pack(side=tk.LEFT, padx=(0, 10))
        
        self.process_button = ttk.Button(buttons_frame,
                                       text="ផ្ទៀងផ្ទាត់ឯកសារ",
                                       style="Primary.TButton",
                                       command=self.process_file,
                                       state=tk.DISABLED)
        self.process_button.pack(side=tk.LEFT)
        
        # Results Frame
        results_frame = ttk.LabelFrame(main_container,
                                     text="លទ្ធផលនៃការត្រួតពិនិត្យ",
                                     style="Custom.TLabelframe",
                                     padding="15")
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        # Output area with custom font and colors
        self.output_area = scrolledtext.ScrolledText(
            results_frame,
            wrap=tk.WORD,
            font=self.khmer_font,
            background="#ffffff",
            foreground="#333333",
            padx=10,
            pady=10
        )
        self.output_area.pack(fill=tk.BOTH, expand=True)
        
        # Status bar
        self.status_var = tk.StringVar(value="រង់ចាំការជ្រើសរើសឯកសារ...")
        status_bar = ttk.Label(main_container,
                             textvariable=self.status_var,
                             font=self.khmer_font,
                             padding=5)
        status_bar.pack(fill=tk.X, pady=(10, 0))

    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="ជ្រើសរើសឯកសារ Excel",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
        )
        
        if file_path:
            # Update file label with filename only
            filename = os.path.basename(file_path)
            self.file_label.config(text=f"ឯកសារដែលបានជ្រើសរើស: {filename}")
            
            # Initialize editor and load workbook
            self.editor = ExcelEditor(file_path)
            load_message = self.editor.load_workbook()
            
            # Update status and enable process button
            self.status_var.set("ឯកសារបានបញ្ចូលរួចរាល់")
            self.process_button.config(state=tk.NORMAL)
            
            messagebox.showinfo("ជោគជ័យ", load_message)
        else:
            self.status_var.set("ការជ្រើសរើសឯកសារត្រូវបានបោះបង់")

    def process_file(self):
        if not self.editor:
            messagebox.showerror("កំហុស", "សូមជ្រើសរើសឯកសារជាមុនសិន")
            return
        
        # Update status
        self.status_var.set("កំពុងដំណើរការ...")
        self.root.update()
        
        # Process the file
        results = self.editor.process_sheets()
        save_message = self.editor.save_workbook()
        
        # Display results
        self.output_area.delete("1.0", tk.END)
        self.output_area.insert(tk.END, results)
        self.output_area.insert(tk.END, f"\n\n{save_message}")
        
        # Update status
        self.status_var.set("ការដំណើរការបានបញ្ចប់")

# Run the Application
if __name__ == "__main__":
    root = ThemedTk(theme="arc")  # You can try different themes like "arc", "equilux", "breeze"
    app = ModernExcelApp(root)
    root.mainloop()
    
# import tkinter as tk
# from tkinter import ttk, filedialog, messagebox, scrolledtext
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# import warnings
# from pathlib import Path
# import os

# # Suppress warnings for unsupported extensions
# warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# class ModernExcelApp:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("Excel Processor")
#         self.root.geometry("900x700")
#         self.editor = None
        
#         # Set the Khmer font and styles
#         self.khmer_font = ("Khmer OS Siemreap", 11)
#         self.setup_styles()
        
#         # Create main container
#         self.create_main_container()
        
#         # Initialize file path
#         self.current_file = None

#     def setup_styles(self):
#         """Setup custom styles for the application"""
#         style = ttk.Style()
        
#         # Configure frame styles
#         style.configure("Main.TFrame", background="#f0f0f0")
#         style.configure("Card.TFrame", background="white", relief="raised")
        
#         # Configure label styles
#         style.configure("Header.TLabel", 
#                        font=("Khmer OS Siemreap", 14, "bold"),
#                        padding=10)
#         style.configure("Status.TLabel", 
#                        font=self.khmer_font,
#                        foreground="#666666")
        
#         # Configure button styles
#         style.configure("Primary.TButton",
#                        font=self.khmer_font,
#                        padding=(20, 10))
#         style.configure("Secondary.TButton",
#                        font=self.khmer_font,
#                        padding=(15, 8))

#     def create_main_container(self):
#         """Create the main container and UI elements"""
#         # Main container
#         self.main_frame = ttk.Frame(self.root, style="Main.TFrame")
#         self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

#         # Header section
#         self.create_header()
        
#         # File selection card
#         self.create_file_selection_card()
        
#         # Status and controls card
#         self.create_status_card()
        
#         # Output area
#         self.create_output_area()

#     def create_header(self):
#         """Create the header section"""
#         header_frame = ttk.Frame(self.main_frame)
#         header_frame.pack(fill=tk.X, pady=(0, 20))
        
#         header_label = ttk.Label(header_frame, 
#                                 text="កម្មវិធីផ្ទៀងផ្ទាត់ឯកសារ Excel",
#                                 style="Header.TLabel")
#         header_label.pack(side=tk.LEFT)

#     def create_file_selection_card(self):
#         """Create the file selection section"""
#         file_frame = ttk.Frame(self.main_frame, style="Card.TFrame")
#         file_frame.pack(fill=tk.X, pady=(0, 20), padx=2, ipady=10)
        
#         # File selection elements
#         select_label = ttk.Label(file_frame,
#                                text="ជ្រើសរើសឯកសារ Excel ដែលត្រូវផ្ទៀងផ្ទាត់",
#                                font=self.khmer_font,
#                                padding=(20, 10))
#         select_label.pack()
        
#         self.file_label = ttk.Label(file_frame,
#                                   text="មិនទាន់មានឯកសារត្រូវបានជ្រើសរើស",
#                                   font=self.khmer_font,
#                                   foreground="#666666",
#                                   padding=(20, 5))
#         self.file_label.pack()
        
#         button_frame = ttk.Frame(file_frame)
#         button_frame.pack(pady=10)
        
#         self.select_button = ttk.Button(button_frame,
#                                       text="ជ្រើសរើសឯកសារ",
#                                       command=self.select_file,
#                                       style="Primary.TButton")
#         self.select_button.pack(side=tk.LEFT, padx=5)

#     def create_status_card(self):
#         """Create the status and controls section"""
#         status_frame = ttk.Frame(self.main_frame, style="Card.TFrame")
#         status_frame.pack(fill=tk.X, pady=(0, 20), padx=2, ipady=10)
        
#         # Status elements
#         status_label = ttk.Label(status_frame,
#                                text="ស្ថានភាព",
#                                font=("Khmer OS Siemreap", 12, "bold"),
#                                padding=(20, 5))
#         status_label.pack()
        
#         self.status_text = ttk.Label(status_frame,
#                                    text="រង់ចាំការជ្រើសរើសឯកសារ",
#                                    style="Status.TLabel",
#                                    padding=(20, 5))
#         self.status_text.pack()
        
#         self.process_button = ttk.Button(status_frame,
#                                        text="ចាប់ផ្តើមផ្ទៀងផ្ទាត់",
#                                        command=self.process_file,
#                                        state=tk.DISABLED,
#                                        style="Primary.TButton")
#         self.process_button.pack(pady=10)

#     def create_output_area(self):
#         """Create the output area"""
#         output_frame = ttk.Frame(self.main_frame, style="Card.TFrame")
#         output_frame.pack(fill=tk.BOTH, expand=True, padx=2)
        
#         # Output header
#         output_header = ttk.Label(output_frame,
#                                 text="លទ្ធផលនៃការផ្ទៀងផ្ទាត់",
#                                 font=("Khmer OS Siemreap", 12, "bold"),
#                                 padding=(20, 10))
#         output_header.pack()
        
#         # Output text area
#         self.output_area = scrolledtext.ScrolledText(
#             output_frame,
#             wrap=tk.WORD,
#             font=self.khmer_font,
#             background="#f8f9fa",
#             pady=10,
#             padx=10
#         )
#         self.output_area.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))

#     def select_file(self):
#         """Handle file selection"""
#         file_path = filedialog.askopenfilename(
#             title="ជ្រើសរើសឯកសារ Excel",
#             filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
#         )
        
#         if file_path:
#             self.current_file = file_path
#             file_name = os.path.basename(file_path)
#             self.file_label.config(
#                 text=f"ឯកសារដែលបានជ្រើសរើស: {file_name}",
#                 foreground="#28a745"
#             )
#             self.editor = ExcelEditor(file_path)
#             load_message = self.editor.load_workbook()
            
#             self.status_text.config(
#                 text="ឯកសារត្រូវបានបញ្ចូលជោគជ័យ",
#                 foreground="#28a745"
#             )
#             self.process_button.config(state=tk.NORMAL)
            
#             messagebox.showinfo("ជោគជ័យ", load_message)
#         else:
#             self.file_label.config(
#                 text="មិនទាន់មានឯកសារត្រូវបានជ្រើសរើស",
#                 foreground="#666666"
#             )

#     def process_file(self):
#         """Handle file processing"""
#         if not self.editor:
#             messagebox.showerror("កំហុស", "សូមជ្រើសរើសឯកសារជាមុនសិន")
#             return
        
#         self.status_text.config(
#             text="កំពុងដំណើរការផ្ទៀងផ្ទាត់...",
#             foreground="#007bff"
#         )
#         self.root.update()
        
#         # Process the file
#         results = self.editor.process_sheets()
#         save_message = self.editor.save_workbook()
        
#         # Update output area
#         self.output_area.delete("1.0", tk.END)
#         self.output_area.insert(tk.END, results)
#         self.output_area.insert(tk.END, f"\n\n{save_message}")
        
#         # Update status
#         self.status_text.config(
#             text="ការផ្ទៀងផ្ទាត់បានបញ្ចប់",
#             foreground="#28a745"
#         )

# # Run the application
# if __name__ == "__main__":
#     root = tk.Tk()
#     app = ModernExcelApp(root)
#     root.mainloop()
    
    
# import tkinter as tk
# from tkinter import ttk, filedialog, messagebox, scrolledtext
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# import warnings
# from ttkthemes import ThemedTk
# import os

# # Suppress warnings for unsupported extensions
# warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# class ExcelEditor:
#     # [Previous ExcelEditor class code remains unchanged]
#     # Note: Keeping all the existing ExcelEditor class functionality

# class ModernExcelApp:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("ប្រព័ន្ធផ្ទៀងផ្ទាត់ឯកសារបំណែងចែកភារកិច្ច")
        
#         # Get screen dimensions
#         screen_width = root.winfo_screenwidth()
#         screen_height = root.winfo_screenheight()
        
#         # Calculate window size and position
#         window_width = int(screen_width * 0.8)
#         window_height = int(screen_height * 0.8)
#         x_position = (screen_width - window_width) // 2
#         y_position = (screen_height - window_height) // 2
        
#         # Set window size and position
#         self.root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
        
#         # Set theme colors
#         self.bg_color = "#f0f0f0"
#         self.accent_color = "#007acc"
#         self.text_color = "#333333"
        
#         # Set fonts
#         self.header_font = ("Khmer OS Siemreap", 16, "bold")
#         self.normal_font = ("Khmer OS Siemreap", 12)
#         self.small_font = ("Khmer OS Siemreap", 10)
        
#         self.editor = None
#         self.create_styles()
#         self.create_widgets()
        
#         # Configure grid weights
#         self.root.grid_rowconfigure(1, weight=1)
#         self.root.grid_columnconfigure(0, weight=1)

#     def create_styles(self):
#         """Create custom styles for widgets"""
#         style = ttk.Style()
        
#         # Configure main theme
#         style.configure("Main.TFrame", background=self.bg_color)
        
#         # Custom button style
#         style.configure("Custom.TButton",
#                        font=self.normal_font,
#                        padding=10)
        
#         # Custom label style
#         style.configure("Custom.TLabel",
#                        font=self.normal_font,
#                        background=self.bg_color,
#                        foreground=self.text_color)
        
#         # Custom labelframe style
#         style.configure("Custom.TLabelframe",
#                        font=self.normal_font,
#                        background=self.bg_color)
        
#         style.configure("Custom.TLabelframe.Label",
#                        font=self.normal_font,
#                        background=self.bg_color,
#                        foreground=self.text_color)

#     def create_widgets(self):
#         """Create and arrange all UI widgets"""
#         # Main container
#         main_container = ttk.Frame(self.root, style="Main.TFrame")
#         main_container.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
#         main_container.grid_columnconfigure(0, weight=1)
#         main_container.grid_rowconfigure(1, weight=1)

#         # Header
#         header_frame = ttk.Frame(main_container, style="Main.TFrame")
#         header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        
#         header_label = ttk.Label(header_frame,
#                                text="ប្រព័ន្ធផ្ទៀងផ្ទាត់ឯកសារបំណែងចែកភារកិច្ច",
#                                font=self.header_font,
#                                style="Custom.TLabel")
#         header_label.pack(pady=10)

#         # File selection frame
#         file_frame = ttk.LabelFrame(main_container,
#                                   text="ជ្រើសរើសឯកសារ",
#                                   style="Custom.TLabelframe")
#         file_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=5)
#         file_frame.grid_columnconfigure(1, weight=1)

#         # File selection widgets
#         self.file_label = ttk.Label(file_frame,
#                                   text="មិនទាន់មានឯកសារត្រូវបានជ្រើសរើស",
#                                   style="Custom.TLabel")
#         self.file_label.grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=5)

#         self.select_button = ttk.Button(file_frame,
#                                       text="ជ្រើសរើសឯកសារ Excel",
#                                       command=self.select_file,
#                                       style="Custom.TButton")
#         self.select_button.grid(row=1, column=0, padx=10, pady=10)

#         self.process_button = ttk.Button(file_frame,
#                                        text="ចាប់ផ្តើមផ្ទៀងផ្ទាត់",
#                                        command=self.process_file,
#                                        state=tk.DISABLED,
#                                        style="Custom.TButton")
#         self.process_button.grid(row=1, column=1, padx=10, pady=10)

#         # Results frame
#         results_frame = ttk.LabelFrame(main_container,
#                                      text="លទ្ធផលនៃការផ្ទៀងផ្ទាត់",
#                                      style="Custom.TLabelframe")
#         results_frame.grid(row=2, column=0, sticky="nsew", padx=5, pady=5)
#         results_frame.grid_columnconfigure(0, weight=1)
#         results_frame.grid_rowconfigure(0, weight=1)

#         # Results text area
#         self.output_area = scrolledtext.ScrolledText(
#             results_frame,
#             wrap=tk.WORD,
#             font=self.normal_font,
#             background="white",
#             foreground=self.text_color
#         )
#         self.output_area.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

#         # Status bar
#         self.status_bar = ttk.Label(main_container,
#                                   text="ត្រៀមរួចរាល់",
#                                   font=self.small_font,
#                                   style="Custom.TLabel")
#         self.status_bar.grid(row=3, column=0, sticky="w", padx=5, pady=5)

#     def select_file(self):
#         """Handle file selection"""
#         file_path = filedialog.askopenfilename(
#             title="ជ្រើសរើសឯកសារ Excel",
#             filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
#         )
        
#         if file_path:
#             self.file_label.config(text=f"ឯកសារដែលបានជ្រើសរើស: {os.path.basename(file_path)}")
#             self.editor = ExcelEditor(file_path)
#             load_message = self.editor.load_workbook()
            
#             if "Error" in load_message:
#                 messagebox.showerror("កំហុស", load_message)
#                 self.status_bar.config(text="មានបញ្ហាក្នុងការបើកឯកសារ")
#             else:
#                 messagebox.showinfo("ជោគជ័យ", "ឯកសារបានបើករួចរាល់")
#                 self.process_button.config(state=tk.NORMAL)
#                 self.status_bar.config(text="ឯកសារត្រូវបានបើករួចរាល់")

#     def process_file(self):
#         """Handle file processing"""
#         if not self.editor:
#             messagebox.showerror("កំហុស", "សូមជ្រើសរើសឯកសារជាមុនសិន")
#             return

#         self.status_bar.config(text="កំពុងដំណើរការ...")
#         self.root.update()

#         try:
#             results = self.editor.process_sheets()
#             save_message = self.editor.save_workbook()

#             self.output_area.delete("1.0", tk.END)
#             self.output_area.insert(tk.END, results)
#             self.output_area.insert(tk.END, f"\n\n{save_message}")
            
#             if "Error" in save_message:
#                 self.status_bar.config(text="មានបញ្ហាក្នុងការរក្សាទុក")
#             else:
#                 self.status_bar.config(text="ការផ្ទៀងផ្ទាត់បានបញ្ចប់")
                
#         except Exception as e:
#             messagebox.showerror("កំហុស", f"មានបញ្ហាក្នុងការដំណើរការ: {str(e)}")
#             self.status_bar.config(text="មានបញ្ហាក្នុងការដំណើរការ")

# # Run the application
# if __name__ == "__main__":
#     root = ThemedTk(theme="arc")  # Using the 'arc' theme for a modern look
#     app = ModernExcelApp(root)
#     root.mainloop()