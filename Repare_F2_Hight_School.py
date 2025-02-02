import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import warnings
from ttkthemes import ThemedTk
import os

# Suppress warnings for unsupported extensions
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class ExcelEditor:

    def __init__(self, file_path):
        """Initialize the editor with the Excel file path."""
        self.file_path = file_path
        self.workbook = None

    def load_workbook(self):
        """Load the workbook from the specified file path."""
        try:
            self.workbook = load_workbook(self.file_path)
            return "ឯកសារបានបញ្ចូលរួចរាល់"
        except Exception as e:
            return f"Error loading workbook: {e}"

    def check_formula_errors(self, wss):
        """
        Check for formula errors in the range A1:FI596.
        Returns a list of cells containing errors.
        """
        ws = wss
        error_cells = []
        error_values = ['#REF!', '#VALUE!', '#NAME?', '#DIV/0!', '#NULL!', '#NUM!', '#N/A']
        
        # Define the range to check (A1:FI596)
        start_row = 1
        end_row = 596
        start_col = 1  # Column A
        end_col = 177  # Column FI (177th column)
        
        # Iterate through the range A1:FI596
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                # Check if cell value is an error
                if isinstance(cell.value, str) and any(error in cell.value for error in error_values):
                    col_letter = get_column_letter(col)
                    error_cells.append(f"Cell {col_letter}{row}")
                    # Highlight error cells in red
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        if error_cells:
            return f"កំហុសរូបមន្តត្រូវបានរកឃើញនៅក្នុង:\n" + "\n".join(error_cells)

    def count_rows_until_condition(self, wss, start_row, column, condition):
        """
        Count rows from a starting row until a condition is met in the specified column.
        Returns the count and whether the condition was met.
        """

        ws = wss
        count = 0
        for row_num in range(start_row, ws.max_row + 1):
            cell_value = ws[f"{column}{row_num}"].value
            count += 1
            if isinstance(cell_value, str) and cell_value.startswith(condition):
                return "", count, True

        return f"មិនបានជួបលក់ខណ្ឌចំណុច {condition} '{ws}'.", count, False

    def check_row(self, wss, start_row, column, condition, num_row, name_blog_row):
        count_message, count, condition_met = self.count_rows_until_condition(wss, start_row, column, condition)
        if count_message:
            return count_message  # Return the error message if there is one

        if condition_met and count != num_row:
            if count < num_row:
                result = num_row-count
                return f"**Sheet '{wss}' បានលុប Row ចំនួន {num_row - count} នៅត្រងចំណុច '{name_blog_row}'.",result
            elif count > num_row:
                result = num_row-count
                return f"**Sheet '{wss}' បានបន្ថែម Row ចំនួន {count - num_row} នៅត្រងចំណុច '{name_blog_row}'.",result

        # Return None if the row count is correct and there's no action needed
        return None,0

    def validate_levels(self, wss, start_row, end_row, level_col, salary_col, certificate_col, day_col, month_col, year_col, gender_col, position_col, add_position_col, grade_col, student_col, subject_1_col, name_blog_row):
        """
        Validate that values in level_col (e.g., 'M') belong to level_salary_a
        and their corresponding values in salary_col (e.g., 'P') exist in the level's set.
        Highlight invalid cells in red.
        """
        # Define a red fill pattern
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        level_salary_a = {
            "ឧត្តម": {
                "ក.3.1", "ក.3.2", "ក.3.3", "ក.3.4",
                "ក.2.1", "ក.2.2", "ក.2.3", "ក.2.4",
                "ក.1.1", "ក.1.2", "ក.1.3", "ក.1.4",
                "ក.1.5", "ក.1.6",
            },
            "មូលដ្ឋាន": {
                "ខ.3.1", "ខ.3.2", "ខ.3.3", "ខ.3.4",
                "ខ.2.1", "ខ.2.2", "ខ.2.3", "ខ.2.4",
                "ខ.1.1", "ខ.1.2", "ខ.1.3", "ខ.1.4",
                "ខ.1.5", "ខ.1.6",
            },
            "បឋម": {
                "គ.1", "គ.2", "គ.3", "គ.4",
                "គ.5", "គ.6", "គ.7", "គ.8",
                "គ.9", "គ.10",
            },
            "មត្តេយ្យ": {
                "គ.1", "គ.2", "គ.3", "គ.4",
                "គ.5", "គ.6", "គ.7", "គ.8",
                "គ.9", "គ.10",
            },
        }

        certificate = {
            "បណ្ឌិត",
            "អនុបណ្ឌិត",
            "បរិញ្ញាបត្រ",
            "មទភ",
            "មបភ",
            "ក្រោម មបភ"
        }

        gander = {
            "ប្រុស",
            "ស្រី"
        }

        position = {
            "នាយក",
            "នាយករង",
            "លេខាធិការ",
            "បណ្ណារក្ស",
            "បេឡា",
            "គណនេយ្យ",
            "ទទួលបន្ទុកយុវជន",
            "បរិវច្ឆការី ",
            "ឆ្មាំ",
            "បម្រើការនៅអង្កការ",
            "បន្តការសិក្សា",
            "កំពុងស្នើលុបឈ្មោះ",
            "សុំចូលនិវត្តមុនអាយុ",
            "ទំនេរគ្មានបៀវត្ស",
            "ក្រៅក្របខណ្ឌដើម",
            "បាត់បង់សម្បទាវិជ្ជាជីវៈ",
            "មានជំងឺរ៉ាំរ៉ៃ",
            "លំហែមាតុភាព",
            "បង្រៀន",
        }

        add_position = {
            "បន្ទុកថ្នាក់",
            "ប្រធានក្រុមបច្ចេកទេស",
            "ប្រ.ក្រុមប.ទ+បន្ទុកថ្នាក់",
            "បង្រៀន",
        }

        grade = {
            "ទី៧",
            "ទី៨",
            "ទី៩",
            "ទី១០",
            "ទី១១ វិ.",
            "ទី១១ ស.",
            "ទី១២ វិ.",
            "ទី១២ ស.",
        }

        subject_1 = {
            "ភាសាខ្មែរ",
            "គណិតវិទ្យា",
            "ភាសាអង់គ្លេស",
            "ភាសាបារាំង",
            "កីឡា",
            "រូបវិទ្យា",
            "គីមីវិទ្យា",
            "ជីវវិទ្យា",
            "ផែនដីវិទ្យា",
            "ប្រវត្តិវិទ្យា",
            "ភូមិវិទ្យា",
            "សីលធម៌-ពលរដ្ឋ",
            "គេហវិទ្យា",
            "សេដ្ឋកិច្ច",
            "ព័ត៌មានវិទ្យា",
            "កសិកម្ម",
            "សិល្បៈ",
            "ដូរ្យតន្រ្តី",
            "នាដសាស្រ្ត",
            "រោងជាង",
            "គ្រប់គ្រងទូទៅ",
            "គ្រប់គ្រងអប់រំ",
            "អេឡិចត្រនិច",
            "អគ្គិសនី",
            "មេកានិច",
            "ភាសារុស្សី",
        }



        ws = wss
        invalid_entries = []

        for row_num in range(start_row, end_row + 1):
            level_value = ws[f"{level_col}{row_num}"].value
            salary_value = ws[f"{salary_col}{row_num}"].value
            certificate_value = ws[f"{certificate_col}{row_num}"].value
            day_value = ws[f"{day_col}{row_num}"].value
            month_value = ws[f"{month_col}{row_num}"].value
            year_value = ws[f"{year_col}{row_num}"].value
            gander_value = ws[f"{gender_col}{row_num}"].value
            position_value = ws[f"{position_col}{row_num}"].value
            add_position_value = ws[f"{add_position_col}{row_num}"].value
            grade_value = ws[f"{grade_col}{row_num}"].value
            student_value = ws[f"{student_col}{row_num}"].value
            subject_1_value = ws[f"{subject_1_col}{row_num}"].value
            
            results = []

            # Check Level of Teacher
            if level_value in level_salary_a:
                if salary_value not in level_salary_a[level_value]:
                    results.append(f"* កាំប្រាក់ពុំត្រឹមត្រូវ '{salary_value}' សម្រាប់ក្របខ័ណ្ឌ '{level_value}' ក្នុង {name_blog_row}")
                    ws[f"{salary_col}{row_num}"].fill = red_fill  # Highlight salary cell in red
                elif salary_value is None:
                    results.append(f"* សូមបញ្ចូលកាំប្រាក់អោយបានត្រឹមត្រូវ {level_value} ក្នុង {name_blog_row}")
                    ws[f"{salary_col}{row_num}"].fill = red_fill  # Highlight salary cell in red

                # Check Certificate
                if certificate_value not in certificate:
                    results.append(f"* បញ្ចូលសញ្ញាបត្រពុំត្រឹមត្រូវ {certificate_value} ក្នុង {name_blog_row}")
                    ws[f"{certificate_col}{row_num}"].fill = red_fill  # Highlight certificate cell in red

                # Try convert and check date of birth
                try:
                    day_value = int(day_value)
                    month_value = int(month_value)
                    year_value = int(year_value)

                    # Check Day of month
                    if day_value is None or not isinstance(day_value, int) or day_value < 1 or day_value > 31:
                        results.append(f"* បញ្ចូលថ្ងៃទីពុំត្រឹមត្រូវ ឬពុំបានបញ្ចូលក្នុង {name_blog_row}")
                        ws[f"{day_col}{row_num}"].fill = red_fill  # Highlight day cell in red

                    # Check Month of Year
                    if month_value is None or not isinstance(month_value, int) or month_value < 1 or month_value > 12:
                        results.append(f"* បញ្ចូលខែពុំត្រឹមត្រូវ ឬពុំបានបញ្ចូលក្នុង {name_blog_row}")
                        ws[f"{month_col}{row_num}"].fill = red_fill  # Highlight month cell in red

                    # Check Year
                    if year_value is None or not isinstance(year_value, int) or year_value < 1964 or year_value > 2006:
                        results.append(f"* បញ្ចូលឆ្នាំពុំត្រឹមត្រូវ ឬពុំបានបញ្ចូលក្នុង {year_value} ក្នុង {name_blog_row}")
                        ws[f"{year_col}{row_num}"].fill = red_fill  # Highlight year cell in red

                except (ValueError, TypeError):
                    results.append(f"* ថ្ងៃខែឆ្នាំកំណើត '{day_value}'| {month_value} | {year_value} មិនមែនជាលេខសូមបញ្ចូលអោយបានត្រឹមត្រូវ.")
                    ws[f"{day_col}{row_num}"].fill = red_fill  # Highlight day cell in red
                    ws[f"{month_col}{row_num}"].fill = red_fill  # Highlight month cell in red
                    ws[f"{year_col}{row_num}"].fill = red_fill  # Highlight year cell in red

                # Check Gender
                if gander_value not in gander:
                    results.append(f"* ភេទមិនត្រឹមត្រូវ ឬពុំបានបញ្ចូល {gander_value} ក្នុង {name_blog_row}")
                    ws[f"{gender_col}{row_num}"].fill = red_fill  # Highlight gender cell in red

                # Check Position
                if position_value not in position:
                    results.append(f"* មុខតំណែងមិនត្រឹមត្រូវ ឬពុំបានបញ្ចូល: {position_value} ក្នុង {name_blog_row}")
                    ws[f"{position_col}{row_num}"].fill = red_fill  # Highlight position cell in red

                # Check Add Position
                if add_position_value in add_position:
                    # check if = បន្ទបថ្នាក់ or ប្រ.ក្រុមប.ទ+បន្ទប់ថ្នាក់
                    if add_position_value == "បន្ទុកថ្នាក់" or add_position_value == "ប្រ.ក្រុមប.ទ+បន្ទុកថ្នាក់":
                        # Check Grade
                        if grade_value not in grade:
                            results.append(f"* មានបញ្ហាចំនួនថ្នាក់ {grade_value} ក្នុង {name_blog_row}")
                            ws[f"{grade_col}{row_num}"].fill = red_fill  # Highlight grade cell in red
                        # Check Total Student
                        if student_value is None or student_value <= 0:
                            results.append(f"* មានបញ្ហាចំនួនសិស្ស {student_value} ក្នុង {name_blog_row}")
                            ws[f"{student_col}{row_num}"].fill = red_fill  # Highlight student cell in red

                elif add_position_value is not None:
                    results.append(f"* ចំណុចភារកិច្ចបន្ថែមមានបញ្ហា '{add_position_value}' ក្នុង {name_blog_row}")
                    ws[f"{add_position_col}{row_num}"].fill = red_fill  # Highlight add position cell in red

                # Check Subject 1
                if name_blog_row == "ក.បុគ្គលិកទីចាត់ការ":
                    if level_value == "បឋម" or level_value == "មត្តេយ្យ":
                        if add_position_value == "បន្ទុកថ្នាក់" or add_position_value == "បង្រៀន":
                            if subject_1_value not in subject_1:
                                results.append(f"* សូមបញ្ចូលមុខវិជ្ជាឱ្យបានត្រឹមត្រូវ {subject_1_value} ក្នុងក្របខ័ណ្ឌ {level_value} នៃ {name_blog_row}")
                                ws[f"{subject_1_col}{row_num}"].fill = red_fill  # Highlight subject cell in red
                    else:
                        if subject_1_value not in subject_1:
                            results.append(f"* សូមបញ្ចូលមុខវិជ្ជាឱ្យបានត្រឹមត្រូវ {subject_1_value} ក្នុងក្របខ័ណ្ឌ {level_value} នៃ {name_blog_row}")
                            ws[f"{subject_1_col}{row_num}"].fill = red_fill  # Highlight subject cell in red
                else:
                    if subject_1_value not in subject_1:
                        results.append(f"* សូមបញ្ចូលមុខវិជ្ជាឱ្យបានត្រឹមត្រូវ {subject_1_value} ក្នុងក្របខ័ណ្ឌ {level_value} នៃ {name_blog_row}")
                        ws[f"{subject_1_col}{row_num}"].fill = red_fill  # Highlight subject cell in red

            elif level_value is not None:
                results.append(f"* បញ្ចូលខុសក្របខ័ណ្ឌអោយបានត្រឹមត្រូវ '{level_value}' ក្នុង {name_blog_row}")
                ws[f"{level_col}{row_num}"].fill = red_fill  # Highlight level cell in red
                
            if results:
                results.insert(0, f"##### សូមពិនិត្យ Row {row_num} : #####")
                invalid_entries.append(results)

        if invalid_entries:
            flattened_entries = [item for sublist in invalid_entries for item in sublist]
            return f"********ការផ្ទៀតផ្ទាត់នៅក្នុងបំណែងចែកភារកិច្ចមានបញ្ហា '{sheet_name}':\n" + "\n".join(map(str, flattened_entries))
    def update_cell(self, wss, cell, value):
        """Update a specific cell with a new value."""
        update_check = []
        ws = wss
        ws[cell] = f"{value}"

        if update_check:
            return f"**Sheet រកពុំឃើញ '{ws}':\n" + "\n".join(update_check)

    def process_sheets(self):
        """Process all sheets starting with 'S' 
        and within the numeric range."""
        cell_var = {
            "DM6": '=COUNTIFS($BB$58:$BB$177,$DM$3,$AC$58:$AC$177,DJ6)+COUNTIFS($BB$179:$BB$328,$DM$3,$AC$179:$AC$328,DJ6)+COUNTIFS($BB$330:$BB$479,$DM$3,$AC$330:$AC$479,DJ6)',
            "DM7": '=COUNTIFS($BB$58:$BB$177,$DM$3,$AC$58:$AC$177,DJ7)+COUNTIFS($BB$179:$BB$328,$DM$3,$AC$179:$AC$328,DJ7)+COUNTIFS($BB$330:$BB$479,$DM$3,$AC$330:$AC$479,DJ7)',
            "DM8": '=COUNTIFS($BB$58:$BB$177,$DM$3,$AC$58:$AC$177,DJ8)+COUNTIFS($BB$179:$BB$328,$DM$3,$AC$179:$AC$328,DJ8)+COUNTIFS($BB$330:$BB$479,$DM$3,$AC$330:$AC$479,DJ8)',
            "DM9": '=COUNTIFS($BB$58:$BB$177,$DM$3,$AC$58:$AC$177,DJ9)+COUNTIFS($BB$179:$BB$328,$DM$3,$AC$179:$AC$328,DJ9)+COUNTIFS($BB$330:$BB$479,$DM$3,$AC$330:$AC$479,DJ9)',
            "DK17": '=SUM(CY16,DA16,DC16,DE16,DG16,DI16)',
            "DK18": '=SUM(CY17,DA17,DC17,DE17,DG17,DI17)',
            "DK19": '=SUM(CY18,DA18,DC18,DE18,DG18,DI18)',
            "DK20": '=SUM(CY19,DA19,DC19,DE19,DG19,DI19)',
            "DK21": '=SUM(CY20,DA20,DC20,DE20,DG20,DI20)',
            "DK22": '=SUM(CY21,DA21,DC21,DE21,DG21,DI21)',
            "DK23": '=SUM(CY22,DA22,DC22,DE22,DG22,DI22)',
            "DK24": '=SUM(CY23,DA23,DC23,DE23,DG23,DI23)',
            "DK26": '=SUM(CY25,DA25,DC25,DE25,DG25,DI25)',
            "J46": '=IF(AY2="",0,IF(AY2="ធម្មតា",0,(AA41-AA38)))',
            "J47": '=IF($AY$2="លំបាក",(AA41-DK20-DK21)*80000,IF($AY$2="ដាច់ស្រយាលប្រភេទ១",(AA41-DK20-DK21)*100000,IF($AY$2="ដាច់ស្រយាលប្រភេទ២",(AA41-DK20-DK21)*120000,IF($AY$2="ធម្មតា",(AA41-DK20-DK21)*0,0))))',
            "AU58": '=IF(AND(Q58="នាយក",$I$16<5),4,IF(AND(Q58="នាយករង",$I$16<8),4,IF(OR(Q58="លេខាធិការ",Q58="បេឡា",Q58="បណ្ណារក្ស",Q58="គណនេយ្យ",Q58="ទទួលបន្ទុកយុវជន",Q58="បរិវច្ឆការី ",Q58="ឆ្មាំ",Q58="លំហែមាតុភាព",Q58="បន្តការសិក្សា",Q58="មានជំងឺរ៉ាំរ៉ៃ",Q58="ក្រៅក្របខណ្ឌដើម"),0,IF(AND(Q58="នាយក",$I$16>=5),0,IF(AND(Q58="នាយករង",$I$16>=8),0,0)))))',
            "BN31": '=BN18+BN20+BN21+BN22+BN23+BN24+BN25+BN26+BN27+BN28+BN29',
            "BX31": '=BX18+BX20+BX21+BX22+BX23+BX24+BX25+BX26+BX27+BX28+BX29'
        }
        results = []
        for sheet in self.workbook.sheetnames:
            if sheet.startswith('S') and sheet[1:].isdigit():
                results.append(f"ដំណើរការផ្ទៀងផ្ទាត់នៅក្នុង: {sheet}")
                ws = self.workbook[sheet]

                error_cells = self.check_formula_errors(ws)
                if isinstance(error_cells, str):  # If it's an error message
                    results.append(error_cells)
                elif error_cells:  # If there are error cells
                    results.append("\nរកឃើញកំហុសរូបមន្តក្នុង:")
                    for error in error_cells:
                        results.append(error)

                # Check rows for different categories
                count = 58
                admin_message = None
                result_admin_row = 0
                result_teacher_high_row = 0
                result_teacher_low_row = 0
                result_contract_row = 0
                admin_message, result_admin_row = self.check_row(ws, count, "A", "ខ", 121, "ក.បុគ្គលិកទីចាត់ការ")
                if result_admin_row != 0:
                    results.append(admin_message)
                    results.append(self.validate_levels(ws, 58, (177-result_admin_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ក.បុគ្គលិកទីចាត់ការ"))
                    count =179-result_admin_row
                    teacher_message_high,result_teacher_high_row = self.check_row(ws, count, "A", "គ", 151, "ខ. គ្រូបង្រៀននៅមធ្យមសិក្សាទុតិយភូមិ")
                    if result_teacher_high_row != 0:
                        results.append(teacher_message_high)
                        results.append(self.validate_levels(ws, count, (328-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ខ. គ្រូបង្រៀននៅមធ្យមសិក្សាទុតិយភូមិ"))
                        count = 330 - result_teacher_high_row - result_admin_row
                        teacher_message_low,result_teacher_low_row = self.check_row(ws, count, "A", "ឃ", 151, "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ")
                        if result_teacher_low_row != 0:
                            results.append(teacher_message_low)
                            results.append(self.validate_levels(ws, count, (479-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ"))
                            count = 481 - result_teacher_low_row - result_teacher_high_row - result_admin_row
                            contract_message, result_contract_row = self.check_row(ws, count, "A", "សរុប", 41, "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា")
                            if contract_message:
                                results.append(contract_message)
                            results.append(self.validate_levels(ws, count, (520-result_admin_row-result_teacher_high_row-result_teacher_low_row-result_contract_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា"))
                        else:
                            results.append(self.validate_levels(ws, count, (479-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ"))
                            count = 481 - result_teacher_high_row - result_admin_row - result_teacher_low_row
                            contract_message, result_contract_row = self.check_row(ws, count, "A", "សរុប", 41, "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា")
                            if contract_message:
                                results.append(contract_message)
                            results.append(self.validate_levels(ws, count, (520-result_admin_row-result_teacher_high_row-result_teacher_low_row-result_contract_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា"))
                    else:
                        results.append(self.validate_levels(ws, count, (328-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ខ. គ្រូបង្រៀននៅមធ្យមសិក្សាទុតិយភូមិ"))
                        count = 330 - result_teacher_high_row - result_admin_row
                        teacher_message_low,result_teacher_low_row = self.check_row(ws, count, "A", "ឃ", 151, "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ")
                        if result_teacher_low_row != 0:
                            results.append(teacher_message_low)
                            results.append(self.validate_levels(ws, count, (479-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ"))
                            count = 481 - result_teacher_low_row - result_teacher_high_row - result_admin_row
                            contract_message, result_contract_row = self.check_row(ws, count, "A", "សរុប", 41, "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា")
                            if contract_message:
                                results.append(contract_message)
                            results.append(self.validate_levels(ws, count, (520-result_admin_row-result_teacher_high_row-result_teacher_low_row-result_contract_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា"))
                        else:
                            results.append(self.validate_levels(ws, count, (479-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ"))
                            count = 481 - result_teacher_high_row - result_admin_row - result_teacher_low_row
                            contract_message, result_contract_row = self.check_row(ws, count, "A", "សរុប", 41, "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា")
                            if contract_message:
                                results.append(contract_message)
                            results.append(self.validate_levels(ws, count, (520-result_admin_row-result_teacher_high_row-result_teacher_low_row-result_contract_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា"))
                else:
                    results.append(self.validate_levels(ws, 58, (177-result_admin_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ក.បុគ្គលិកទីចាត់ការ"))
                    count =179-result_admin_row
                    teacher_message_high,result_teacher_high_row = self.check_row(ws, count, "A", "គ", 151, "ខ. គ្រូបង្រៀននៅមធ្យមសិក្សាទុតិយភូមិ")
                    if result_teacher_high_row != 0:
                        results.append(teacher_message_high)
                        results.append(self.validate_levels(ws, count, (328-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ខ. គ្រូបង្រៀននៅមធ្យមសិក្សាទុតិយភូមិ"))
                        count = 330 - result_teacher_high_row - result_admin_row
                        teacher_message_low,result_teacher_low_row = self.check_row(ws, count, "A", "ឃ", 151, "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ")
                        if result_teacher_low_row != 0:
                            results.append(teacher_message_low)
                            results.append(self.validate_levels(ws, count, (479-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ"))
                            count = 481 - result_teacher_low_row - result_teacher_high_row - result_admin_row
                            contract_message, result_contract_row = self.check_row(ws, count, "A", "សរុប", 41, "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា")
                            if contract_message:
                                results.append(contract_message)
                            results.append(self.validate_levels(ws, count, (520-result_admin_row-result_teacher_high_row-result_teacher_low_row-result_contract_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា"))
                        else:
                            results.append(self.validate_levels(ws, count, (479-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ"))
                            count = 481 - result_teacher_high_row - result_admin_row - result_teacher_low_row
                            contract_message, result_contract_row = self.check_row(ws, count, "A", "សរុប", 41, "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា")
                            if contract_message:
                                results.append(contract_message)
                            results.append(self.validate_levels(ws, count, (520-result_admin_row-result_teacher_high_row-result_teacher_low_row-result_contract_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា"))
                    else:
                        results.append(self.validate_levels(ws, count, (328-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ខ. គ្រូបង្រៀននៅមធ្យមសិក្សាទុតិយភូមិ"))
                        count = 330 - result_teacher_high_row - result_admin_row
                        teacher_message_low,result_teacher_low_row = self.check_row(ws, count, "A", "ឃ", 151, "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ")
                        if result_teacher_low_row != 0:
                            results.append(teacher_message_low)
                            results.append(self.validate_levels(ws, count, (479-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ"))
                            count = 481 - result_teacher_low_row - result_teacher_high_row - result_admin_row
                            contract_message, result_contract_row = self.check_row(ws, count, "A", "សរុប", 41, "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា")
                            if contract_message:
                                results.append(contract_message)
                            results.append(self.validate_levels(ws, count, (520-result_admin_row-result_teacher_high_row-result_teacher_low_row-result_contract_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា"))
                        else:
                            results.append(self.validate_levels(ws, count, (479-result_admin_row-result_teacher_high_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "គ. គ្រូបង្រៀននៅមធ្យមសិក្សាបឋមភូមិ"))
                            count = 481 - result_teacher_high_row - result_admin_row - result_teacher_low_row
                            contract_message, result_contract_row = self.check_row(ws, count, "A", "សរុប", 41, "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា")
                            if contract_message:
                                results.append(contract_message)
                            results.append(self.validate_levels(ws, count, (520-result_admin_row-result_teacher_high_row-result_teacher_low_row-result_contract_row), "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "ឃ. គ្រូខ្ចី មន្ត្រីជាប់កិច្ចសន្យា និងគ្រូជាប់កិច្ចសន្យា"))
                
                ### Update Cell That

                for Key in cell_var.keys():
                    for result in [
                        self.update_cell(ws, Key, cell_var[Key])
                    ]:
                        if result:  # Only append non-None results
                            results.append(result)

            else:
                results.append(f"មិនមែនជា Sheet សាលារំលង: {sheet}")

        # Filter out None values before joining
        return "\n".join(filter(None, results))

    def save_workbook(self):
        """Save the workbook back to the file."""
        try:
            self.workbook.save(self.file_path)
            return "កិច្ចការបានរក្សាទុករួចរាល់។"
        except Exception as e:
            return f"មានបញ្ហាក្នុងការរក្សារទុក ឬ Excel កំពុងបើក: {e}"


class ModernExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Processor")
        self.root.geometry("800x700")
        self.editor = None
        
        # Set the Khmer font and styles
        self.khmer_font = ("Khmer OS Siemreap", 11)
        self.khmer_font_bold = ("Khmer OS Siemreap", 11, "bold")
        
        # Configure the theme
        self.configure_styles()
        
        # Create UI elements
        self.create_widgets()
        
    def configure_styles(self):
        # Configure custom styles
        style = ttk.Style()
        
        # Configure frame styles
        style.configure("Custom.TFrame", background="#f0f0f0")
        
        # Configure button styles
        style.configure("Primary.TButton",
                       font=self.khmer_font,
                       padding=10)
        
        # Configure label styles
        style.configure("Header.TLabel",
                       font=self.khmer_font_bold,
                       padding=5)
        
        style.configure("Info.TLabel",
                       font=self.khmer_font,
                       padding=5)
        
        # Configure labelframe styles
        style.configure("Custom.TLabelframe",
                       font=self.khmer_font_bold,
                       padding=10)
        style.configure("Custom.TLabelframe.Label",
                       font=self.khmer_font_bold)

    def create_widgets(self):
        # Main container with padding
        main_container = ttk.Frame(self.root, style="Custom.TFrame", padding="20")
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header_frame = ttk.Frame(main_container)
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        header_label = ttk.Label(header_frame,
                                text="Script ត្រួតពិនិត្យទិន្នន័យគ្រូបង្រៀន F1 F2 នៅមធ្យមសិក្សាបឋមភូមិ និងទុតិយភូមិ",
                                style="Header.TLabel",
                                font=("Khmer OS Siemreap", 16, "bold"))
        header_label.pack()
        
        # File Selection Frame
        file_frame = ttk.LabelFrame(main_container,
                                  text="ជ្រើសរើសឯកសារ",
                                  style="Custom.TLabelframe",
                                  padding="15")
        file_frame.pack(fill=tk.X, pady=(0, 20))
        
        # File info container
        file_info_frame = ttk.Frame(file_frame)
        file_info_frame.pack(fill=tk.X, pady=(0, 10))
        
        # File icon label (you can replace with actual icon)
        file_icon_label = ttk.Label(file_info_frame, text="📄", font=("", 24))
        file_icon_label.pack(side=tk.LEFT, padx=(0, 10))
        
        self.file_label = ttk.Label(file_info_frame,
                                  text="ឯកសារមិនទាន់បានជ្រើសរើស",
                                  style="Info.TLabel",
                                  wraplength=600)
        self.file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Buttons container
        buttons_frame = ttk.Frame(file_frame)
        buttons_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.select_button = ttk.Button(buttons_frame,
                                      text="ជ្រើសរើសឯកសារ",
                                      style="Primary.TButton",
                                      command=self.select_file)
        self.select_button.pack(side=tk.LEFT, padx=(0, 10))
        
        self.process_button = ttk.Button(buttons_frame,
                                       text="ផ្ទៀងផ្ទាត់ឯកសារ",
                                       style="Primary.TButton",
                                       command=self.process_file,
                                       state=tk.DISABLED)
        self.process_button.pack(side=tk.LEFT)
        
        # Results Frame
        results_frame = ttk.LabelFrame(main_container,
                                     text="លទ្ធផលនៃការត្រួតពិនិត្យ",
                                     style="Custom.TLabelframe",
                                     padding="15")
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        # Output area with custom font and colors
        self.output_area = scrolledtext.ScrolledText(
            results_frame,
            wrap=tk.WORD,
            font=self.khmer_font,
            background="#ffffff",
            foreground="#333333",
            padx=10,
            pady=10
        )
        self.output_area.pack(fill=tk.BOTH, expand=True)
        
        # Status bar
        self.status_var = tk.StringVar(value="រង់ចាំការជ្រើសរើសឯកសារ...")
        status_bar = ttk.Label(main_container,
                             textvariable=self.status_var,
                             font=self.khmer_font,
                             padding=5)
        status_bar.pack(fill=tk.X, pady=(10, 0))

    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="ជ្រើសរើសឯកសារ Excel",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
        )
        
        if file_path:
            # Update file label with filename only
            filename = os.path.basename(file_path)
            self.file_label.config(text=f"ឯកសារដែលបានជ្រើសរើស: {filename}")
            
            # Initialize editor and load workbook
            self.editor = ExcelEditor(file_path)
            load_message = self.editor.load_workbook()
            
            # Update status and enable process button
            self.status_var.set("ឯកសារបានបញ្ចូលរួចរាល់")
            self.process_button.config(state=tk.NORMAL)
            
            messagebox.showinfo("ជោគជ័យ", load_message)
        else:
            self.status_var.set("ការជ្រើសរើសឯកសារត្រូវបានបោះបង់")

    def process_file(self):
        if not self.editor:
            messagebox.showerror("កំហុស", "សូមជ្រើសរើសឯកសារជាមុនសិន")
            return
        
        # Update status
        self.status_var.set("កំពុងដំណើរការ...")
        self.root.update()
        
        # Process the file
        results = self.editor.process_sheets()
        save_message = self.editor.save_workbook()
        
        # Display results
        self.output_area.delete("1.0", tk.END)
        self.output_area.insert(tk.END, results)
        self.output_area.insert(tk.END, f"\n\n{save_message}")
        
        # Update status
        self.status_var.set("ការដំណើរការបានបញ្ចប់")

# Run the Application
if __name__ == "__main__":
    root = ThemedTk(theme="arc")  # You can try different themes like "arc", "equilux", "breeze"
    app = ModernExcelApp(root)
    root.mainloop()