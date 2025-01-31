import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from openpyxl import load_workbook
import warnings

# Suppress warnings for unsupported extensions
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class ExcelEditor:
    def __init__(self, file_path):
        """Initialize the editor with the Excel file path."""
        self.file_path = file_path
        self.workbook = None

    def load_workbook(self):
        """Load the workbook from the specified file path."""
        try:
            self.workbook = load_workbook(self.file_path)
            return "Workbook loaded successfully."
        except Exception as e:
            return f"Error loading workbook: {e}"

    def count_rows_until_condition(self, sheet_name, start_row, column, condition):
        """
        Count rows from a starting row until a condition is met in the specified column.
        Returns the count and whether the condition was met.
        """
        if sheet_name not in self.workbook.sheetnames:
            return f"Sheet '{sheet_name}' not found.", 0, False

        ws = self.workbook[sheet_name]
        count = 0

        for row_num in range(start_row, ws.max_row + 1):
            cell_value = ws[f"{column}{row_num}"].value
            count += 1
            if isinstance(cell_value, str) and cell_value.startswith(condition):
                return "", count, True

        return f"Condition not met in sheet '{sheet_name}'.", count, False

    def check_row(self, sheet, start_row, column, condition, num_row, name_blog_row):
        count_message, count, condition_met = self.count_rows_until_condition(sheet, start_row, column, condition)
        if count_message:
            return count_message  # Return the error message if there is one

        if condition_met and count != num_row:
            if count < num_row:
                return f"Sheet '{sheet}' had {num_row - count} rows deleted in '{name_blog_row}'."
            elif count > num_row:
                return f"Sheet '{sheet}' had {count - num_row} rows inserted in '{name_blog_row}'."

        # Return None if the row count is correct and there's no action needed
        return None

    def validate_levels(self, sheet_name, start_row, end_row, level_col, salary_col, certificate_col, day_col, month_col, year_col, gender_col, position_col, add_position_col, grade_col, student_col, subject_1_col, name_blog_row):
        """
        Validate that values in level_col (e.g., 'M') belong to level_salary_a
        and their corresponding values in salary_col (e.g., 'P') exist in the level's set.
        """
        level_salary_a = {
            "ឧត្តម": {
                "ក.3.1", "ក.3.2", "ក.3.3", "ក.3.4",
                "ក.2.1", "ក.2.2", "ក.2.3", "ក.2.4",
                "ក.1.1", "ក.1.2", "ក.1.3", "ក.1.4",
                "ក.1.5", "ក.1.6",
            },
            "មូលដ្ឋាន": {
                "ខ.3.1", "ខ.3.2", "ខ.3.3", "ខ.3.4",
                "ខ.2.1", "ខ.2.2", "ខ.2.3", "ខ.2.4",
                "ខ.1.1", "ខ.1.2", "ខ.1.3", "ខ.1.4",
                "ខ.1.5", "ខ.1.6",
            },
            "បឋម": {
                "គ.1", "គ.2", "គ.3", "គ.4",
                "គ.5", "គ.6", "គ.7", "គ.8",
                "គ.9", "គ.10",
            },
            "មត្តេយ្យ": {
                "គ.1", "គ.2", "គ.3", "គ.4",
                "គ.5", "គ.6", "គ.7", "គ.8",
                "គ.9", "គ.10",
            },
        }

        certificate = {
            "បណ្ឌិត",
            "អនុបណ្ឌិត",
            "បរិញ្ញាបត្រ",
            "មទភ",
            "មបភ",
            "ក្រោម មបភ"
        }

        gander = {
            "ប្រុស",
            "ស្រី"
        }

        position = {
            "នាយក",
            "នាយករង",
            "លេខាធិការ",
            "បណ្ណារក្ស",
            "បេឡា",
            "គណនេយ្យ",
            "ទទួលបន្ទុកយុវជន",
            "បរិវច្ឆការី ",
            "ឆ្មាំ",
            "បម្រើការនៅអង្កការ",
            "បន្តការសិក្សា",
            "កំពុងស្នើលុបឈ្មោះ",
            "សុំចូលនិវត្តមុនអាយុ",
            "ទំនេរគ្មានបៀវត្ស",
            "ក្រៅក្របខណ្ឌដើម",
            "បាត់បង់សម្បទាវិជ្ជាជីវៈ",
            "មានជំងឺរ៉ាំរ៉ៃ",
            "លំហែមាតុភាព",
            "បង្រៀន",
        }

        add_position = {
            "បន្ទុកថ្នាក់",
            "ប្រធានក្រុមបច្ចេកទេស",
            "ប្រ.ក្រុមប.ទ+បន្ទុកថ្នាក់",
            "បង្រៀន",
        }

        grade = {
            "ទី៧",
            "ទី៨",
            "ទី៩",
            "ទី១០",
            "ទី១១ វិ.",
            "ទី១១ ស.",
            "ទី១២ វិ.",
            "ទី១២ ស.",
        }

        subject_1 = {
            "ភាសាខ្មែរ",
            "គណិតវិទ្យា",
            "ភាសាអង់គ្លេស",
            "ភាសាបារាំង",
            "កីឡា",
            "រូបវិទ្យា",
            "គីមីវិទ្យា",
            "ជីវវិទ្យា",
            "ផែនដីវិទ្យា",
            "ប្រវត្តិវិទ្យា",
            "ភូមិវិទ្យា",
            "សីលធម៌-ពលរដ្ឋ",
            "គេហវិទ្យា",
            "សេដ្ឋកិច្ច",
            "ព័ត៌មានវិទ្យា",
            "កសិកម្ម",
            "សិល្បៈ",
            "ដូរ្យតន្រ្តី",
            "នាដសាស្រ្ត",
            "រោងជាង",
            "គ្រប់គ្រងទូទៅ",
            "គ្រប់គ្រងអប់រំ",
            "អេឡិចត្រនិច",
            "អគ្គិសនី",
            "មេកានិច",
            "ភាសារុស្សី",
        }

        if sheet_name not in self.workbook.sheetnames:
            return f"Sheet '{sheet_name}' not found."

        ws = self.workbook[sheet_name]
        invalid_entries = []

        for row_num in range(start_row, end_row + 1):
            level_value = ws[f"{level_col}{row_num}"].value
            salary_value = ws[f"{salary_col}{row_num}"].value
            certificate_value = ws[f"{certificate_col}{row_num}"].value
            day_value = ws[f"{day_col}{row_num}"].value
            month_value = ws[f"{month_col}{row_num}"].value
            year_value = ws[f"{year_col}{row_num}"].value
            gander_value = ws[f"{gender_col}{row_num}"].value
            position_value = ws[f"{position_col}{row_num}"].value
            add_position_value = ws[f"{add_position_col}{row_num}"].value
            grade_value = ws[f"{grade_col}{row_num}"].value
            student_value = ws[f"{student_col}{row_num}"].value
            subject_1_value = ws[f"{subject_1_col}{row_num}"].value

            # Check Level of Teacher
            if level_value in level_salary_a:
                if salary_value not in level_salary_a[level_value]:
                    invalid_entries.append(f"Row {row_num}: Invalid salary '{salary_value}' for level '{level_value}' in {name_blog_row}")
                elif salary_value is None:
                    invalid_entries.append(f"Row {row_num}: Salary Level not input in Level {level_value} in {name_blog_row}")

                # Check Certificate
                if certificate_value not in certificate:
                    invalid_entries.append(f"Row {row_num}: Invalid Certificate in {certificate_value} in {name_blog_row}")

                # Try convert and check date of birth
                try:
                    day_value = int(day_value)
                    month_value = int(month_value)
                    year_value = int(year_value)

                    # Check Day of month
                    if day_value is None or not isinstance(day_value, int) or day_value < 1 or day_value > 31:
                        invalid_entries.append(f"Row {row_num}: Invalid Day or Forgot Day in {day_value} in {name_blog_row}")

                    # Check Month of Year
                    if month_value is None or not isinstance(month_value, int) or month_value < 1 or month_value > 12:
                        invalid_entries.append(f"Row {row_num}: Invalid Month or Forgot Month in {month_value} in {name_blog_row}")

                    # Check Year
                    if year_value is None or not isinstance(year_value, int) or year_value < 1964 or year_value > 2006:
                        invalid_entries.append(f"Row {row_num}: Invalid Year or Forgot Year in {year_value} in {name_blog_row}")

                except (ValueError, TypeError):
                    invalid_entries.append(f"Row {row_num}: Data of birth '{day_value}'| {month_value} | {year_value} is not a valid integer.")

                # Check Gender
                if gander_value not in gander:
                    invalid_entries.append(f"Row {row_num}: Invalid Gander or Forgot in {gander_value} in {name_blog_row}")

                # Check Position
                if position_value not in position:
                    invalid_entries.append(f"Row {row_num}: Invalid Position or Forgot in {position_value} in {name_blog_row}")

                # Check Add Position
                if add_position_value in add_position:
                    # check if = បន្ទបថ្នាក់ or ប្រ.ក្រុមប.ទ+បន្ទប់ថ្នាក់
                    if add_position_value == "បន្ទុកថ្នាក់" or add_position_value == "ប្រ.ក្រុមប.ទ+បន្ទុកថ្នាក់":
                        # Check Grade
                        if grade_value not in grade:
                            invalid_entries.append(f"Row {row_num}: Invalid Grade or Forgot in {grade_value} in {name_blog_row}")
                        # Check Total Student
                        if student_value is None or student_value <= 0:
                            invalid_entries.append(f"Row {row_num}: Invalid Student or Forgot in {student_value} in {name_blog_row}")

                elif add_position_value is not None:
                    invalid_entries.append(f"Row {row_num}: Unknown Add Position '{add_position_value}' in {name_blog_row}")

                # Check Subject 1
                if subject_1_value not in subject_1:
                    invalid_entries.append(f"Row {row_num}: Unknown Subject 1 in {subject_1_value} in {name_blog_row}")

            elif level_value is not None:
                invalid_entries.append(f"Row {row_num}: Unknown level '{level_value}' in {name_blog_row}")

        if invalid_entries:
            return f"Validation errors in sheet '{sheet_name}':\n" + "\n".join(invalid_entries)

    def update_cell(self, sheet_name, cell, value):
        """Update a specific cell with a new value."""
        update_check = []
        if sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            ws[cell] = f"={value}"
            update_check.append(f"Updated {cell} in sheet '{sheet_name}'")
        else:
            update_check.append(f"Sheet '{sheet_name}' not found.")

        if update_check:
            return f"Updated in sheet '{sheet_name}':\n" + "\n".join(update_check)

    def process_sheets(self):
        """Process all sheets starting with 'S' 
        and within the numeric range."""
        cell_var = {
            "DM6": '=COUNTIFS($BB$58:$BB$177,$DM$3,$AC$58:$AC$177,DJ6)+COUNTIFS($BB$179:$BB$328,$DM$3,$AC$179:$AC$328,DJ6)+COUNTIFS($BB$330:$BB$479,$DM$3,$AC$330:$AC$479,DJ6)',
            "DM7": '=COUNTIFS($BB$58:$BB$177,$DM$3,$AC$58:$AC$177,DJ7)+COUNTIFS($BB$179:$BB$328,$DM$3,$AC$179:$AC$328,DJ7)+COUNTIFS($BB$330:$BB$479,$DM$3,$AC$330:$AC$479,DJ7)',
            "DM8": '=COUNTIFS($BB$58:$BB$177,$DM$3,$AC$58:$AC$177,DJ8)+COUNTIFS($BB$179:$BB$328,$DM$3,$AC$179:$AC$328,DJ8)+COUNTIFS($BB$330:$BB$479,$DM$3,$AC$330:$AC$479,DJ8)',
            "DM9": '=COUNTIFS($BB$58:$BB$177,$DM$3,$AC$58:$AC$177,DJ9)+COUNTIFS($BB$179:$BB$328,$DM$3,$AC$179:$AC$328,DJ9)+COUNTIFS($BB$330:$BB$479,$DM$3,$AC$330:$AC$479,DJ9)',
            "DK17": '=SUM(CY16,DA16,DC16,DE16,DG16,DI16)',
            "DK18": '=SUM(CY17,DA17,DC17,DE17,DG17,DI17)',
            "DK19": '=SUM(CY18,DA18,DC18,DE18,DG18,DI18)',
            "DK20": '=SUM(CY19,DA19,DC19,DE19,DG19,DI19)',
            "DK21": '=SUM(CY20,DA20,DC20,DE20,DG20,DI20)',
            "DK22": '=SUM(CY21,DA21,DC21,DE21,DG21,DI21)',
            "DK23": '=SUM(CY22,DA22,DC22,DE22,DG22,DI22)',
            "DK24": '=SUM(CY23,DA23,DC23,DE23,DG23,DI23)',
            "DK26": '=SUM(CY25,DA25,DC25,DE25,DG25,DI25)',
            "J46": '=IF(AY2="",0,IF(AY2="ធម្មតា",0,(AA41-AA38)))',
            "J47": '=IF($AY$2="លំបាក",(AA41-DK21-DK22)*80000,IF($AY$2="ដាច់ស្រយាលប្រភេទ១",(AA41-DK21-DK22)*100000,IF($AY$2="ដាច់ស្រយាលប្រភេទ២",(AA41-DK21-DK22)*120000,IF($AY$2="ធម្មតា",(AA41-DK21-DK22)*0,0))))',
        }
        results = []
        for sheet in self.workbook.sheetnames:
            if sheet.startswith('S') and sheet[1:].isdigit() and 1 <= int(sheet[1:]) <= 30:
                results.append(f"Processing sheet: {sheet}")

                # Check rows for different categories
                for result in [
                    self.check_row(sheet, 58, "A", "ខ", 121, "Office"),
                    self.check_row(sheet, 179, "A", "គ", 151, "High School"),
                    self.check_row(sheet, 330, "A", "ឃ", 151, "Secondary School"),
                    self.check_row(sheet, 481, "A", "សរុប", 41, "Contract")
                ]:
                    if result:  # Only append non-None results
                        results.append(result)

                ### Update Cell That

                for Key in cell_var.keys():
                    for result in [
                        self.update_cell(sheet, Key, cell_var[Key])
                    ]:
                        if result:  # Only append non-None results
                            results.append(result)

                # Additional checks for levels in column M and P
                for result in [
                    self.validate_levels(sheet, 58, 177, "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "Office"),
                    self.validate_levels(sheet, 179, 328, "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "High School"),
                    self.validate_levels(sheet, 330, 479, "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "Secondary School"),
                    self.validate_levels(sheet, 481, 221, "M", "P", "AC", "I", "J", "K", "L", "Q", "T", "V", "W", "Y", "Contract School")
                ]:
                    if result:  # Only append non-None results
                        results.append(result)
            else:
                results.append(f"Skipped sheet: {sheet}")

        # Filter out None values before joining
        return "\n".join(filter(None, results))

    def save_workbook(self):
        """Save the workbook back to the file."""
        try:
            self.workbook.save(self.file_path)
            return "Workbook saved successfully."
        except Exception as e:
            return f"Error saving workbook: {e}"


# GUI Implementation
class ExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Processor")
        self.root.geometry("600x500")
        self.editor = None

        # Create UI elements
        self.create_widgets()

    def create_widgets(self):
        # Main Frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # File Selection
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.pack(fill=tk.X, pady=5)

        self.file_label = ttk.Label(file_frame, text="No file selected", wraplength=500, justify="left")
        self.file_label.pack(pady=10)

        self.select_button = ttk.Button(file_frame, text="Select Excel File", command=self.select_file)
        self.select_button.pack(pady=5)

        self.process_button = ttk.Button(file_frame, text="Process File", command=self.process_file, state=tk.DISABLED)
        self.process_button.pack(pady=5)

        # Output Area
        output_frame = ttk.LabelFrame(main_frame, text="Output", padding="10")
        output_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        self.output_area = scrolledtext.ScrolledText(output_frame, wrap="word", height=20, width=70)
        self.output_area.pack(fill=tk.BOTH, expand=True)

    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
        )
        if file_path:
            self.file_label.config(text=f"Selected File: {file_path}")
            self.editor = ExcelEditor(file_path)
            load_message = self.editor.load_workbook()
            messagebox.showinfo("File Loaded", load_message)
            self.process_button.config(state=tk.NORMAL)
        else:
            self.file_label.config(text="No file selected")

    def process_file(self):
        if not self.editor:
            messagebox.showerror("Error", "Please select a file first.")
            return

        results = self.editor.process_sheets()
        save_message = self.editor.save_workbook()

        # Display results in the output area
        self.output_area.delete("1.0", tk.END)
        self.output_area.insert(tk.END, results)
        self.output_area.insert(tk.END, f"\n\n{save_message}")


# Run the Application
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelApp(root)
    root.mainloop()