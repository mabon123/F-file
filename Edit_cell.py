from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

class ExcelEditor:
    def __init__(self, file_path):
        """Initialize the editor with the Excel file path."""
        self.file_path = file_path
        self.workbook = None

    def load_workbook(self):
        """Load the workbook from the specified file path."""
        try:
            self.workbook = load_workbook(self.file_path)
            print("Workbook loaded successfully.")
        except Exception as e:
            print(f"Error loading workbook: {e}")

    def update_cell(self, sheet_name, cell, value):
        """Update a specific cell with a new value."""
        if sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            ws[cell].value = value
            print(f"Updated {cell} in sheet '{sheet_name}' with value: {value}")
        else:
            print(f"Sheet '{sheet_name}' not found.")

    def count_rows_until_condition(self, sheet_name, start_row, column, condition):
        """
        Count rows from a starting row until a condition is met in the specified column.
        Returns the count and whether the condition was met.
        """
        if sheet_name not in self.workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found.")
            return 0, False

        ws = self.workbook[sheet_name]
        count = 0

        for row_num in range(start_row, ws.max_row + 1):
            cell_value = ws[f"{column}{row_num}"].value
            count += 1
            if isinstance(cell_value, str) and cell_value.startswith(condition):
                return count, True

        print(f"Condition not met in sheet '{sheet_name}'.")
        return count, False
    
    def check_row(self, sheet, start_row,column, condition, num_row, name_blog_row):
        count, condition_met = self.count_rows_until_condition(sheet, start_row, column, condition)
        if condition_met and count != num_row:
            if count < num_row:
                print(f"Sheet '{sheet}' had {num_row - count} been Delete in '{name_blog_row}'.")
            elif count > num_row:
                print(f"Sheet'{sheet}' had '{count-num_row}' been Insert.'{name_blog_row}'")
            else:
                print(f"Sheet '{sheet}' need Check again.")
        else:
            print(f"Row count is correct in sheet '{sheet}'.")

    def process_sheets(self):
        """Process all sheets starting with 'S' and within the numeric range."""
        for sheet in self.workbook.sheetnames:
            if sheet.startswith('S') and sheet[1:].isdigit() and 1 <= int(sheet[1:]) <= 30:
                print(f"Processing sheet: {sheet}")
                # check_row Office
                self.check_row(
                    sheet,
                    start_row=58,
                    column="A",
                    condition="ខ",
                    num_row=121,
                    name_blog_row="Office"
                )
                
                # check_row High School
                self.check_row(
                    sheet,
                    start_row=179,
                    column="A",
                    condition="គ",
                    num_row=151,
                    name_blog_row="High School"
                )
                
                # check_row Secondary School
                self.check_row(
                    sheet,
                    start_row=330,
                    column="A",
                    condition="ឃ",
                    num_row=151,
                    name_blog_row="Secondary School"
                )
                # check_row contract
                self.check_row(
                    sheet,
                    start_row=481,
                    column="A",
                    condition="សរុប",
                    num_row=41,
                    name_blog_row="Contract"
                )
                
            else:
                print(f"Skipped sheet: {sheet}")
                
    # def count_row(self, ):
        

    def save_workbook(self):
        """Save the workbook back to the file."""
        try:
            self.workbook.save(self.file_path)
            print("Workbook saved successfully.")
        except Exception as e:
            print(f"Error saving workbook: {e}")
            



# Example usage
if __name__ == "__main__":
    #Open a file dialog for the user to select the file
    root = tk.Tk()
    root.withdraw() #Hide the root window
    work_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel file", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    if not work_path:
        print("No file selected. Exiting")
    else:
        editor = ExcelEditor(work_path)
        editor.load_workbook()
        #Check Row in School Office
        editor.process_sheets()
        editor.save_workbook()
